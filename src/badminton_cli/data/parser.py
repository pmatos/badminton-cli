"""Parse Excel ranking files into Player objects."""

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..models.player import Discipline, Player, RankingWeek

COLUMN_MAP: dict[str, str] = {
    "GS": "gender",
    "DIS": "discipline",
    "Ranglistenplatz": "rank",
    "FRang": "federation_rank",
    "Nachname": "last_name",
    "Vorname": "first_name",
    "SpielerID": "player_id",
    "GJahr": "birth_year",
    "AKL1": "age_class_1",
    "AKL2": "age_class_2",
    "Points": "points",
    "Turniere": "tournaments",
    "Verein": "club",
    "Bezirk": "district",
}


def parse_excel(
    path: Path,
    ranking_week: RankingWeek | None = None,
) -> list[Player]:
    """Parse an Excel ranking file into Player objects.

    Args:
        path: Path to the Excel file.
        ranking_week: Optional RankingWeek to associate with players.

    Returns:
        List of Player objects.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws: Worksheet | None = wb.active

    if ws is None:
        wb.close()
        return []

    rows: list[tuple[Any, ...]] = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []

    headers = rows[0]
    col_indices: dict[str, int] = {}
    for i, header in enumerate(headers):
        header_str = str(header) if header is not None else ""
        if header_str in COLUMN_MAP:
            col_indices[COLUMN_MAP[header_str]] = i

    if ranking_week is None:
        ranking_week = extract_week_from_filename(path.name)

    week_label: str | None = ranking_week.label if ranking_week else None

    players: list[Player] = []
    for row in rows[1:]:
        if not row or row[0] is None:
            continue

        try:
            discipline_str = str(row[col_indices["discipline"]])
            if discipline_str not in [d.value for d in Discipline]:
                continue

            points_raw = row[col_indices["points"]]
            points: float = float(points_raw) / 1000.0 if points_raw else 0.0

            player = Player(
                discipline=Discipline(discipline_str),
                rank=int(row[col_indices["rank"]] or 0),
                federation_rank=int(row[col_indices["federation_rank"]] or 0),
                last_name=str(row[col_indices["last_name"]] or ""),
                first_name=str(row[col_indices["first_name"]] or ""),
                gender=str(row[col_indices["gender"]] or ""),
                player_id=str(row[col_indices["player_id"]] or ""),
                birth_year=int(row[col_indices["birth_year"]] or 0),
                age_class_1=str(row[col_indices["age_class_1"]] or ""),
                age_class_2=str(row[col_indices["age_class_2"]] or ""),
                points=points,
                tournaments=int(row[col_indices["tournaments"]] or 0),
                club=str(row[col_indices["club"]] or ""),
                district=str(row[col_indices["district"]] or ""),
                ranking_week=week_label,
            )
            players.append(player)
        except (ValueError, KeyError, TypeError):
            continue

    wb.close()
    return players


def extract_week_from_filename(filename: str) -> RankingWeek | None:
    """Extract ranking week info from filename.

    Args:
        filename: Filename like 'Ranking_2026_KW02.xlsx'

    Returns:
        RankingWeek object or None if pattern doesn't match.
    """
    pattern = r"Ranking_(\d{4})_KW(\d{2})\.xlsx"
    match = re.match(pattern, filename)
    if match:
        year = int(match.group(1))
        week = int(match.group(2))
        return RankingWeek(year=year, week=week)
    return None
